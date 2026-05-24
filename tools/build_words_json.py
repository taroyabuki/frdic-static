#!/usr/bin/env python3
import argparse
import json
import os
import sqlite3
from datetime import datetime, timezone
from pathlib import Path


def normalize_search_key(value: str) -> str:
    s = (value or "").strip().lower()
    table = str.maketrans({
        "à": "a", "á": "a", "â": "a", "ä": "a", "ã": "a", "å": "a",
        "ç": "c",
        "è": "e", "é": "e", "ê": "e", "ë": "e",
        "ì": "i", "í": "i", "î": "i", "ï": "i",
        "ñ": "n",
        "ò": "o", "ó": "o", "ô": "o", "ö": "o", "õ": "o",
        "ù": "u", "ú": "u", "û": "u", "ü": "u",
        "ý": "y", "ÿ": "y",
        "œ": "oe", "æ": "ae",
    })
    s = s.translate(table)
    for ch in [" ", "\t", "\n", "\r", "-", "’", "'", "."]:
        s = s.replace(ch, "")
    return "".join(ch for ch in s if ("a" <= ch <= "z") or ("0" <= ch <= "9"))


def is_verb_category(category: str) -> bool:
    c = category or ""
    if "動詞" in c:
        return True
    lower = c.casefold()
    return "verb" in lower or "verbe" in lower


def build_present_conjugation(infinitive: str):
    w = (infinitive or "").strip()
    if not w:
        return []
    lower = w.casefold()
    irregular = {
        "être": ["suis", "es", "est", "sommes", "êtes", "sont"],
        "etre": ["suis", "es", "est", "sommes", "êtes", "sont"],
        "avoir": ["ai", "as", "a", "avons", "avez", "ont"],
        "aller": ["vais", "vas", "va", "allons", "allez", "vont"],
        "faire": ["fais", "fais", "fait", "faisons", "faites", "font"],
    }
    if lower in irregular:
        return irregular[lower]
    if lower.endswith("er"):
        stem = w[:-2]
        nous_stem = stem
        if lower.endswith("ger"):
            nous_stem = stem + "e"
        elif lower.endswith("cer") and stem:
            nous_stem = stem[:-1] + "ç"
        return [stem + "e", stem + "es", stem + "e", nous_stem + "ons", stem + "ez", stem + "ent"]
    if lower.endswith("ir"):
        stem = w[:-2]
        return [stem + "is", stem + "is", stem + "it", stem + "issons", stem + "issez", stem + "issent"]
    if lower.endswith("re"):
        stem = w[:-2]
        return [stem + "s", stem + "s", stem, stem + "ons", stem + "ez", stem + "ent"]
    return []


def build_imparfait_conjugation(infinitive: str, present_forms):
    if len(present_forms) != 6:
        return []
    lower = (infinitive or "").strip().casefold()
    if lower in ("être", "etre"):
        stem = "ét"
    else:
        nous = present_forms[3]
        if not nous.endswith("ons"):
            return []
        stem = nous[:-3]
    return [stem + "ais", stem + "ais", stem + "ait", stem + "ions", stem + "iez", stem + "aient"]


def build_future_simple_conjugation(infinitive: str):
    w = (infinitive or "").strip()
    if not w:
        return []
    lower = w.casefold()
    irregular = {
        "être": "ser", "etre": "ser", "avoir": "aur", "aller": "ir", "faire": "fer",
        "venir": "viendr", "voir": "verr", "pouvoir": "pourr", "vouloir": "voudr",
        "devoir": "devr", "savoir": "saur",
    }
    if lower in irregular:
        stem = irregular[lower]
    elif lower.endswith("re"):
        stem = w[:-1]
    else:
        stem = w
    return [stem + "ai", stem + "as", stem + "a", stem + "ons", stem + "ez", stem + "ont"]


def is_etre_aux_verb(lower_infinitive: str) -> bool:
    return lower_infinitive in {
        "aller", "venir", "arriver", "partir", "entrer", "sortir", "monter", "descendre",
        "naître", "naitre", "mourir", "tomber", "rester", "retourner", "devenir", "revenir",
    }


def past_participle(lower_infinitive: str) -> str:
    irregular = {
        "être": "été", "etre": "été", "avoir": "eu", "faire": "fait", "dire": "dit",
        "prendre": "pris", "mettre": "mis", "voir": "vu", "pouvoir": "pu",
        "vouloir": "voulu", "devoir": "dû", "savoir": "su", "venir": "venu",
        "tenir": "tenu", "aller": "allé", "naître": "né", "naitre": "né", "mourir": "mort",
    }
    if lower_infinitive in irregular:
        return irregular[lower_infinitive]
    if lower_infinitive.endswith("er"):
        return lower_infinitive[:-2] + "é"
    if lower_infinitive.endswith("ir"):
        return lower_infinitive[:-2] + "i"
    if lower_infinitive.endswith("re"):
        return lower_infinitive[:-2] + "u"
    return ""


def build_passe_compose_conjugation(infinitive: str):
    lower = (infinitive or "").strip().casefold()
    if not lower:
        return []
    pp = past_participle(lower)
    if not pp:
        return []
    aux = ["suis", "es", "est", "sommes", "êtes", "sont"] if is_etre_aux_verb(lower) else ["ai", "as", "a", "avons", "avez", "ont"]
    return [f"{aux[0]} {pp}", f"{aux[1]} {pp}", f"{aux[2]} {pp}", f"{aux[3]} {pp}", f"{aux[4]} {pp}", f"{aux[5]} {pp}"]


def generated_conjugations(word: str, category: str):
    if not is_verb_category(category):
        return empty_conjugations()
    present = build_present_conjugation(word)
    if len(present) != 6:
        return empty_conjugations()
    return {
        "present": present,
        "imparfait": build_imparfait_conjugation(word, present),
        "future": build_future_simple_conjugation(word),
        "passe_compose": build_passe_compose_conjugation(word),
    }


def empty_conjugations():
    return {
        "present": [],
        "imparfait": [],
        "future": [],
        "passe_compose": [],
    }


def parse_conjugation_json(value):
    if not value:
        return []
    try:
        parsed = json.loads(value)
    except json.JSONDecodeError:
        return []
    if not isinstance(parsed, list) or len(parsed) != 6:
        return []
    if not all(isinstance(item, str) for item in parsed):
        return []
    return parsed


def db_columns(cur) -> set[str]:
    return {row[1] for row in cur.execute("PRAGMA table_info(words)")}


def rows_from_db(path: Path):
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cols = db_columns(cur)
    optional = {
        "example_fr": "'' AS example_fr",
        "example_ja": "'' AS example_ja",
        "conj_present": "'' AS conj_present",
        "conj_imparfait": "'' AS conj_imparfait",
        "conj_future": "'' AS conj_future",
        "conj_passe_compose": "'' AS conj_passe_compose",
    }
    select_cols = [
        "id",
        "word",
        "word_norm" if "word_norm" in cols else "'' AS word_norm",
        "COALESCE(category, '') AS category",
        "meaning",
    ]
    for col, fallback in optional.items():
        select_cols.append(f"COALESCE({col}, '') AS {col}" if col in cols else fallback)

    rows = []
    try:
        for row in cur.execute(f"SELECT {', '.join(select_cols)} FROM words ORDER BY word"):
            word = clean_cell(row["word"])
            word_norm = clean_cell(row["word_norm"]) or normalize_search_key(word)
            rows.append({
                "id": int(row["id"]),
                "word": word,
                "word_norm": word_norm,
                "category": clean_cell(row["category"]),
                "meaning": clean_cell(row["meaning"]),
                "example_fr": clean_cell(row["example_fr"]),
                "example_ja": clean_cell(row["example_ja"]),
                "conjugations": {
                    "present": parse_conjugation_json(row["conj_present"]),
                    "imparfait": parse_conjugation_json(row["conj_imparfait"]),
                    "future": parse_conjugation_json(row["conj_future"]),
                    "passe_compose": parse_conjugation_json(row["conj_passe_compose"]),
                },
            })
    finally:
        conn.close()
    return rows


def rows_from_xlsx(path: Path):
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    next_id = 1
    for row in ws.iter_rows(values_only=True):
        word = clean_cell(row[0] if len(row) >= 1 else "")
        category = clean_cell(row[1] if len(row) >= 2 else "")
        meaning = clean_cell(row[2] if len(row) >= 3 else "")
        example_fr = clean_cell(row[3] if len(row) >= 4 else "")
        example_ja = clean_cell(row[4] if len(row) >= 5 else "")
        if not word and not category and not meaning and not example_fr and not example_ja:
            continue
        if not word or not meaning:
            continue
        rows.append({
            "id": next_id,
            "word": word,
            "word_norm": normalize_search_key(word),
            "category": category,
            "meaning": meaning,
            "example_fr": example_fr,
            "example_ja": example_ja,
            "conjugations": generated_conjugations(word, category),
        })
        next_id += 1
    return rows


def clean_cell(value) -> str:
    return "" if value is None else str(value).strip()


def filter_rows(rows, initials: str | None):
    if not initials:
        return rows
    allowed = {ch for ch in normalize_search_key(initials)}
    if not allowed:
        return rows
    return [row for row in rows if row["word_norm"][:1] in allowed]


def write_json(path: Path, payload: dict, pretty: bool):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        if pretty:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        else:
            json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
        f.write("\n")


def build_payload(rows, source_kind: str, source_name: str, initials: str | None):
    return {
        "schema_version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "source": {
            "kind": source_kind,
            "name": source_name,
        },
        "filter": {
            "initials": initials or None,
            "normalized_initials": bool(initials),
        },
        "words": rows,
    }


def main():
    parser = argparse.ArgumentParser(description="Build browser-loadable frdic JSON data.")
    inputs = parser.add_mutually_exclusive_group(required=True)
    inputs.add_argument("--input-db", type=Path, help="SQLite words.db input path")
    inputs.add_argument("--input-xlsx", type=Path, help="Excel input path, headerless A-E format")
    parser.add_argument("--output", type=Path, default=Path("docs/words.json"), help="Output JSON path")
    parser.add_argument("--initials", default="", help="Keep words whose normalized word starts with any of these characters")
    parser.add_argument("--pretty", action="store_true", help="Pretty-print JSON")
    args = parser.parse_args()

    if args.input_db:
        rows = rows_from_db(args.input_db)
        source_kind = "sqlite"
        source_name = os.fspath(args.input_db)
    else:
        rows = rows_from_xlsx(args.input_xlsx)
        source_kind = "xlsx"
        source_name = os.fspath(args.input_xlsx)

    rows = filter_rows(rows, args.initials)
    payload = build_payload(rows, source_kind, source_name, args.initials or None)
    write_json(args.output, payload, args.pretty)

    print(f"source={source_kind}")
    print(f"rows_written={len(rows)}")
    print(f"output={args.output}")


if __name__ == "__main__":
    main()
